#!/usr/bin/env python3
"""
MCP Server for converting Excel files to Markdown.

Provides tools to list sheets and convert Excel/CSV data to Markdown tables
via Streamable HTTP transport.
"""

import csv
import json
import os
import re
from pathlib import Path
from typing import Annotated, Optional

import openpyxl
from mcp.server.fastmcp import FastMCP
from pydantic import Field

PORT = int(os.getenv("PORT", "6000"))

mcp = FastMCP("excel_to_md_mcp")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _read_xlsx(path: Path) -> dict[str, list[list]]:
    """Return {sheet_name: [[row values...]]} for every sheet in an xlsx/xlsm file."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: dict[str, list[list]] = {}
    for name in wb.sheetnames:
        ws = wb[name]
        sheets[name] = [
            [cell.value if cell.value is not None else "" for cell in row]
            for row in ws.iter_rows()
        ]
    wb.close()
    return sheets


def _read_csv(path: Path) -> dict[str, list[list]]:
    """Return a single-sheet dict for CSV files."""
    with open(path, newline="", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))
    return {path.stem: rows}


def _load_file(file_path: str) -> dict[str, list[list]]:
    """Dispatch to the correct reader based on file extension."""
    path = Path(file_path).resolve()
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    ext = path.suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_xlsx(path)
    if ext == ".csv":
        return _read_csv(path)
    raise ValueError(f"Unsupported file type '{ext}'. Supported: .xlsx, .xlsm, .csv")


def _rows_to_markdown(sheet_name: str, rows: list[list]) -> str:
    """Convert a list of rows to a Markdown table section."""
    if not rows:
        return f"## {sheet_name}\n\n_Empty sheet_\n"

    col_count = max((len(r) for r in rows), default=0)

    def normalize(row: list) -> str:
        cells = [str(row[i] if i < len(row) else "").replace("|", "\\|").replace("\n", " ") for i in range(col_count)]
        return "| " + " | ".join(cells) + " |"

    header = normalize(rows[0])
    separator = "| " + " | ".join(["---"] * col_count) + " |"
    body = "\n".join(normalize(r) for r in rows[1:])

    return f"## {sheet_name}\n\n{header}\n{separator}\n{body}\n"


def _parse_markdown_tables(markdown_content: str) -> dict[str, list[list]]:
    """Parse Markdown content into {sheet_name: [[row values...]]} structure.
    
    Expects format:
    ## Sheet Name
    | Header1 | Header2 |
    | --- | --- |
    | Value1 | Value2 |
    """
    sheets: dict[str, list[list]] = {}
    
    # Split by level-2 headings
    sections = re.split(r'^## (.+)$', markdown_content, flags=re.MULTILINE)
    
    # sections[0] is content before first heading (usually empty)
    # sections[1] is first sheet name, sections[2] is its content, etc.
    for i in range(1, len(sections), 2):
        if i + 1 >= len(sections):
            break
            
        sheet_name = sections[i].strip()
        content = sections[i + 1].strip()
        
        # Skip empty sheets
        if not content or content == "_Empty sheet_":
            sheets[sheet_name] = []
            continue
        
        # Extract table rows
        rows = []
        lines = content.split('\n')
        
        for line in lines:
            line = line.strip()
            # Skip empty lines and separator lines
            if not line or re.match(r'^\|[\s\-:|]+\|$', line):
                continue
            
            # Parse table row
            if line.startswith('|') and line.endswith('|'):
                # Remove leading/trailing pipes and split
                cells = line[1:-1].split('|')
                # Unescape pipes and trim whitespace
                cells = [cell.strip().replace('\\|', '|') for cell in cells]
                rows.append(cells)
        
        if rows:
            sheets[sheet_name] = rows
    
    return sheets


def _write_excel(file_path: str, sheets_data: dict[str, list[list]]) -> None:
    """Write sheets data to an Excel file."""
    path = Path(file_path).resolve()
    
    # Create parent directory if needed
    path.parent.mkdir(parents=True, exist_ok=True)
    
    wb = openpyxl.Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    for sheet_name, rows in sheets_data.items():
        ws = wb.create_sheet(title=sheet_name)
        for row_data in rows:
            ws.append(row_data)
    
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# Tools  (flat params — no Pydantic wrapper, works with all MCP clients)
# ---------------------------------------------------------------------------

FilePath = Annotated[str, Field(description="Absolute or relative path to the Excel/CSV file (e.g. 'C:/data/report.xlsx')", min_length=1)]
SheetsList = Annotated[Optional[list[str]], Field(default=None, description="Sheet names to convert. Omit or pass null to convert all sheets.")]


@mcp.tool(
    name="excel_list_sheets",
    annotations={
        "title": "List Excel Sheets",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
async def excel_list_sheets(file_path: FilePath) -> str:
    """List all sheet names in an Excel or CSV file (.xlsx, .xlsm, .csv).

    Returns JSON: {"file": str, "sheet_count": int, "sheets": [str]}
    On error returns: "Error: <message>"
    """
    try:
        data = _load_file(file_path)
        result = {
            "file": str(Path(file_path).resolve()),
            "sheet_count": len(data),
            "sheets": list(data.keys()),
        }
        return json.dumps(result, indent=2, ensure_ascii=False)
    except (FileNotFoundError, ValueError) as e:
        return f"Error: {e}"
    except Exception as e:
        return f"Error: Unexpected error - {type(e).__name__}: {e}"


@mcp.tool(
    name="excel_convert_to_markdown",
    annotations={
        "title": "Convert Excel to Markdown",
        "readOnlyHint": True,
        "destructiveHint": False,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
async def excel_convert_to_markdown(file_path: FilePath, sheets: SheetsList = None) -> str:
    """Convert an Excel or CSV file to Markdown tables.

    Each sheet becomes a level-2 heading followed by a Markdown table.
    The first row is treated as the table header.

    Examples:
        - All sheets:   file_path="report.xlsx"
        - One sheet:    file_path="report.xlsx", sheets=["Summary"]
        - CSV:          file_path="data.csv"

    On error returns: "Error: <message>"
    """
    try:
        all_sheets = _load_file(file_path)

        target_names = sheets if sheets else list(all_sheets.keys())
        missing = [s for s in target_names if s not in all_sheets]
        if missing:
            available = ", ".join(f'"{n}"' for n in all_sheets.keys())
            return f"Error: Sheet(s) not found: {missing}. Available: [{available}]"

        parts = [_rows_to_markdown(name, all_sheets[name]) for name in target_names]
        return "\n".join(parts)
    except (FileNotFoundError, ValueError) as e:
        return f"Error: {e}"
    except Exception as e:
        return f"Error: Unexpected error - {type(e).__name__}: {e}"


MarkdownContent = Annotated[str, Field(description="Markdown content with tables to convert. Expected format: ## Sheet Name followed by Markdown tables.", min_length=1)]
OutputFilePath = Annotated[str, Field(description="Path where the Excel file will be created (e.g. 'C:/data/output.xlsx'). Will create parent directories if needed.", min_length=1)]


@mcp.tool(
    name="markdown_convert_to_excel",
    annotations={
        "title": "Convert Markdown to Excel",
        "readOnlyHint": False,
        "destructiveHint": True,
        "idempotentHint": True,
        "openWorldHint": False,
    },
)
async def markdown_convert_to_excel(markdown_content: MarkdownContent, output_file_path: OutputFilePath) -> str:
    """Convert Markdown tables to an Excel file.

    Parses Markdown content with level-2 headings (##) as sheet names,
    followed by Markdown tables. Creates an Excel file with multiple sheets.

    Expected Markdown format:
        ## Sheet1
        | Header1 | Header2 |
        | --- | --- |
        | Value1 | Value2 |

        ## Sheet2
        | ColA | ColB |
        | --- | --- |
        | Data1 | Data2 |

    Returns JSON: {"file": str, "sheet_count": int, "sheets": [str]}
    On error returns: "Error: <message>"

    Examples:
        - Basic: markdown_content="## Data\n| A | B |\n| --- | --- |\n| 1 | 2 |", output_file_path="output.xlsx"
    """
    try:
        # Parse markdown into sheets
        sheets_data = _parse_markdown_tables(markdown_content)
        
        if not sheets_data:
            return "Error: No valid Markdown tables found. Expected format: ## Sheet Name followed by table."
        
        # Validate output path
        output_path = Path(output_file_path).resolve()
        ext = output_path.suffix.lower()
        if ext not in {".xlsx", ".xlsm"}:
            return f"Error: Output file must have .xlsx or .xlsm extension, got '{ext}'"
        
        # Write to Excel
        _write_excel(str(output_path), sheets_data)
        
        result = {
            "file": str(output_path),
            "sheet_count": len(sheets_data),
            "sheets": list(sheets_data.keys()),
        }
        return json.dumps(result, indent=2, ensure_ascii=False)
        
    except PermissionError as e:
        return f"Error: Permission denied writing to '{output_file_path}'. Check file permissions or if file is open."
    except Exception as e:
        return f"Error: Unexpected error - {type(e).__name__}: {e}"


# ---------------------------------------------------------------------------
# ASGI middleware: patch Accept header for clients that don't send it correctly
# (e.g. CodeVista). FastMCP requires both application/json and text/event-stream.
# ---------------------------------------------------------------------------

class FixAcceptMiddleware:
    def __init__(self, app):
        self.app = app

    async def __call__(self, scope, receive, send):
        if scope["type"] == "http":
            headers = {k.lower(): v for k, v in scope["headers"]}
            accept = headers.get(b"accept", b"").decode()
            if "application/json" not in accept or "text/event-stream" not in accept:
                patched = [(k, v) for k, v in scope["headers"] if k.lower() != b"accept"]
                patched.append((b"accept", b"application/json, text/event-stream"))
                scope = {**scope, "headers": patched}
        await self.app(scope, receive, send)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import uvicorn

    print(f"Excel->Markdown MCP server starting on http://localhost:{PORT}/mcp")
    app = FixAcceptMiddleware(mcp.streamable_http_app())
    uvicorn.run(app, host="0.0.0.0", port=PORT)
